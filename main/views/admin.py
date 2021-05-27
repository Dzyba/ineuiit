from django.views.generic import View
from django.shortcuts import redirect, render
from django.templatetags.static import static
from django.http import HttpResponseRedirect, HttpResponse
from main.models import Menu, ScheduleGroup, ScheduleDay, ScheduleTimeSlot
from django.contrib import messages
from django.urls import reverse
from main.forms import AdminPushForm
from webpush import send_group_notification
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.writer.excel import save_virtual_workbook
from datetime import datetime


class AdminScheduleXLSXView(View):
    def get(self, request, *args, **kwargs):
        group = ScheduleGroup.objects.get(slug=kwargs['slug'])
        schedule = group.get_schedule_dict()

        wb = Workbook()
        ws = wb.active
        ws.title = group.name
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30

        # Стили
        success_fill = PatternFill(start_color='cfe3d7', end_color='cfe3d7', fill_type='solid')
        info_fill = PatternFill(start_color='e2eff7', end_color='e2eff7', fill_type='solid')

        # Данные
        day_row = 1
        for day in schedule:
            ws.cell(row=day_row, column=1).value = day['day'].get_day_display()
            ws.cell(row=day_row, column=1).font = Font(bold=True)

            ws.cell(row=day_row, column=2).value = 'Время'
            ws.cell(row=day_row, column=2).font = Font(bold=True)

            ws.cell(row=day_row, column=3).value = 'ЧС'
            ws.cell(row=day_row, column=3).font = Font(bold=True)
            ws.cell(row=day_row, column=3).fill = success_fill

            ws.cell(row=day_row, column=4).value = 'ЗН'
            ws.cell(row=day_row, column=4).font = Font(bold=True)
            ws.cell(row=day_row, column=4).fill = info_fill

            timeslot_row = 1
            for timeslot in day['timeslots']:
                ws.cell(row=day_row+timeslot_row, column=2).value = timeslot.time
                if timeslot.is_whole:
                    ws.merge_cells(
                        start_row=day_row + timeslot_row,
                        start_column=3,
                        end_row=day_row + timeslot_row,
                        end_column=4
                    )
                    ws.cell(row=day_row + timeslot_row, column=3).value = timeslot.pair_numerator
                else:
                    if timeslot.pair_numerator:
                        ws.cell(row=day_row + timeslot_row, column=3).value = timeslot.pair_numerator
                    if timeslot.pair_denominator:
                        ws.cell(row=day_row + timeslot_row, column=4).value = timeslot.pair_denominator
                timeslot_row += 1
            day_row += day['rowspan'] + 2

        response = HttpResponse(
            content=save_virtual_workbook(wb),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': 'attachment; filename=schedule_%s_%s.xlsx' % (group.slug, datetime.now().strftime('%Y-%m-%d_%H_%M_%S'))
            }
        )
        return response

class AdminPushView(View):
    template_name = 'admin/push.html'

    def _get_context(self, request, *args, **kwargs):
        context = {}
        return context

    def get(self, request, *args, **kwargs):
        context = self._get_context(request, *args, **kwargs)

        ref = request.META.get('HTTP_REFERER', '')
        context['form'] = AdminPushForm(initial={'path': ref})

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        if 'apply' in request.POST:
            push_form = AdminPushForm(request.POST)
            if push_form.is_valid():
                header = push_form.cleaned_data['header']
                text = push_form.cleaned_data['text']

                payload = {
                    'head': header,
                    'body': text,
                    'icon': static('main/img/logo-square.png'),
                    'url': reverse('main:schedules')
                }
                send_group_notification(group_name='all', payload=payload, ttl=1000)

                messages.add_message(request, messages.INFO, 'Уведомление отправлено', extra_tags='', fail_silently=False)
                return HttpResponseRedirect(reverse('admin:main_schedulegroup_changelist'))

        push_form = AdminPushForm(initial={'path': request.get_full_path()})
        return render(request, 'admin/push.html', context={'form': push_form})


# Вспомогательные функции сообщений и редиректа
def _message(request, message, status=messages.INFO):
    messages.add_message(request, status, message, extra_tags='', fail_silently=False)
    return redirect('admin:main_menu_changelist')

def _error(request, message='Не удалось изменить порядок'):
    return _message(request, message, messages.ERROR)


class AdminMenuOrderUpView(View):
    def get(self, request, *args, **kwargs):
        if not 'id' in kwargs:
            return _error(request)
        try:
            menu = Menu.objects.get(id=kwargs['id'])
        except:
            return _error(request)
        if not menu.parent:
            return _error(request)

        if menu.order_in_parent > 1:
            menu.order_up()
        else:
            return _error(request, 'Порядок уже первый в своём пункте, возможно вам надо изменить родительский пункт')

        return _message(request, 'Порядок изменён вверх')

class AdminMenuOrderDownView(View):
    def get(self, request, *args, **kwargs):
        if not 'id' in kwargs:
            return _error(request)
        try:
            menu = Menu.objects.get(id=kwargs['id'])
        except:
            return _error(request)
        if not menu.parent:
            return _error(request)

        if not menu.order_in_parent_is_last:
            menu.order_down()
        else:
            return _error(request, 'Порядок уже последний в своём пункте, возможно вам надо изменить родительский пункт')

        return _message(request, 'Порядок изменён вниз')

